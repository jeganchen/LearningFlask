from flask import Flask, render_template, request
from HelloFlask import app
from datetime import datetime
from werkzeug.utils import secure_filename
import openpyxl
import os
import pyodbc


@app.route('/')
@app.route('/home')
def home():
    now = datetime.now()
    format_date = now.strftime("%A, %d %B, %Y at %X")
    title = "Hello Flask"
    message = "Hello Flask!"
    content = "on " + format_date
    return render_template("index.html", title=title, message=message, content=content)

@app.route('/about')
def about():
    return render_template("about.html",
                           title='About Flask',
                           content='all about flask')

@app.route('/upload', methods=['GET', 'POST'])
def upload():
    if request.method == 'GET':
        return render_template("upload.html",
                               title="upload file")
    elif request.method == 'POST':
        if 'fileUpload' not in request.files:
            flash('No file part')
            return redirect(request.url)
        file = request.files['fileUpload']
        UPLOAD_FOLDER = './HelloFlask/tempfiles/{}.xlsx'.format(str(int(datetime.now().timestamp()*100)))   # 第一种方法
        ROOT_FOLDER = os.path.join(os.getcwd(), UPLOAD_FOLDER)  # 整合绝对路径
        file.save(ROOT_FOLDER)
      #  retrun render_template("upload.html", result=insert_table(ROOT_FOLDER))
     #   return load_excel(ROOT_FOLDER) #
        return render_template("upload.html", result=insert_table(ROOT_FOLDER))

def load_excel(ROOT_FOLDER):
    wb = openpyxl.load_workbook(ROOT_FOLDER)
    ws = wb.active
    rown = ws.max_row
    coln = ws.max_column
    content = '<table name=data border="1"> '
    for row in ws.iter_rows(max_row=rown, max_col=coln):
        content += ' <tr> '
        for cell in row:
            content += ' <td> '+ str(cell.value) + '</td>'
        content += '</tr>'
    content +='</table>'
    return content

def insert_table(ROOT_FOLDER):
    connection = pyodbc.connect(
        driver='{iSeries Access ODBC Driver}',
        system='gitdev',
        uid='chanfree',
        pwd='mattel01')
    c1=connection.cursor()

    wb = openpyxl.load_workbook(ROOT_FOLDER)
    ws = wb.active
    rown = ws.max_row
    coln = ws.max_column
    content = ' '
    for row in ws.iter_rows(min_row=2, max_row=rown, max_col=coln):
        for cell in row:
            if content != ' ':
                content +=","
            content += "'"+ str(cell.value) + "'"
        sql = 'insert into devjegan.TRUMSPTM values ({})'.format(content)
        content = ' '
        print(sql)
        c1.execute(sql)
    c1.execute('select COMPTR , TOYNTR , ORNOTR,FOBPTR,SELYTR,EFDTTR,ETDTTR,SDOCTR,URGTTR from devjegan.TRUMSPTM ')
    connection.commit()
    return c1
    



